import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Border, Side

def atoms_for_hbond(protein):
    def check_hbond(atom):
        if atom.name[0] in ['H', 'O', 'N']:
            if atom.residue.res_name != 'HOH':
                if len(atom.residue.bonds[atom.name]) != 1:
                    return False
            return True
        return False

    atoms_in_hbonds = []
    hbonds_donors = []
    atoms_in_qq = []
    qq_negative = []
    min_coord = np.array([float('inf'), float('inf'), float('inf')])
    max_coord = np.array([-float('inf'), -float('inf'), -float('inf')])
    for v in protein.residues.values():
        for atom in v.atoms.values():
            protein.atoms.update({atom.atom_id: atom})
            # skip backbone
            if 'C' in atom.residue.bonds[atom.name] or 'N' in atom.residue.bonds[atom.name]:
                continue
            # check hbonds
            if check_hbond(atom):
                atoms_in_hbonds.append(atom)
                if atom.name[0] in ['O', 'N']:
                    hbonds_donors.append(atom)

            # charge - charge
            if atom.residue.charge > 0:
                # check NH2-
                if atom.name[0] in ['N']:
                    hydrogens = list(filter(lambda s: s.startswith('H'), atom.residue.bonds[atom.name]))
                    if len(hydrogens) == 2:
                        atoms_in_qq.append(atom)
                        atoms_in_qq.extend(list(map(atom.residue.atoms.get, hydrogens, [None]*len(hydrogens))))
            elif atom.residue.charge < 0:
                if atom.name[0] in ['C']:
                    atom_o = list(filter(lambda s: s.startswith('O'), atom.residue.bonds[atom.name]))
                    if len(atom_o) == 2:
                        atoms_in_qq.extend(list(map(atom.residue.atoms.get, atom_o, [None]*len(atom_o))))
                        qq_negative.append(atom)
        coords = [atom.coordinates for atom in v.atoms.values()]
        min_coord = np.min(coords + [min_coord], axis=0)
        max_coord = np.max(coords + [max_coord], axis=0)
    return atoms_in_hbonds, hbonds_donors, atoms_in_qq, qq_negative, min_coord, max_coord


def format_interaction(j, idx):
    """
    Форматирует информацию о взаимодействии, выводя сначала атом с меньшим residue.res_id.

    Args:
    j: Объект, содержащий информацию о взаимодействии и атомах (atom1, atom2).

    Returns:
    Отформатированная строка.
    """
    if j.atom1.residue.res_id == idx:
        return f'{j.atom1.residue.res_id} - {j.atom1.name}, {j.atom2.residue.res_id} - {j.atom2.name}'
    else:
        return f'{j.atom2.residue.res_id} - {j.atom2.name}, {j.atom1.residue.res_id} - {j.atom1.name}'


def get_inteructions(protein, pdb_name, hbond_protein_dict, electrostatic_protein_dict):
    sorted_items = sorted(protein.residues.items())
    for idx, res in sorted_items:
        if not (any([atom.hbond for atom in res.atoms.values()]) or any([atom.electrostatic for atom in res.atoms.values()])):
            continue
        for a in res.atoms.values():
            if a.hbond:
                hbond_protein_dict[f'{format_interaction(a.hbond, idx)}'] = hbond_protein_dict.get(
                    f'{format_interaction(a.hbond, idx)}', []) + [pdb_name]
            if a.electrostatic:
                electrostatic_protein_dict[f'{format_interaction(a.electrostatic, idx)}'] = electrostatic_protein_dict.get(
                    f'{format_interaction(a.electrostatic, idx)}', []) + [pdb_name]
    return hbond_protein_dict, electrostatic_protein_dict


def create_excel(my_dict, columns, file_name):
    # Создаем новую книгу Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Создаем стили для границ
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Записываем заголовки столбцов
    sheet.cell(row=1, column=1, value="Key") # записываем заголовок первого столбца "Key"
    for col_num, col_name in enumerate(columns, start=2):
        cell = sheet.cell(row=1, column=col_num, value=col_name)
        cell.border = thin_border  # Добавляем границы к заголовкам

    # Записываем данные
    red_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid") # Создаём стиль заливки красным
    for row_num, (key, value_list) in enumerate(my_dict, start=2):
        cell = sheet.cell(row=row_num, column=1, value=key) # записываем ключи в первый столбец
        cell.border = thin_border  # Добавляем границы к заголовкам

        for col_num, col_name in enumerate(columns, start=2):
            cell = sheet.cell(row=row_num, column=col_num)  # Получаем ячейку
            if col_name in value_list:
                cell.fill = red_fill  # окрашиваем ячейку в красный
            else:
                cell.value = '-'
            cell.border = thin_border  # Добавляем границы

    # Сохраняем файл
    workbook.save(f"{file_name}.xlsx")

